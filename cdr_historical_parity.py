"""Canonical multiset parity and conservative legacy semantic recovery."""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
import json
import re
from typing import Any, Iterable, Iterator, Mapping, Sequence

from cdr_historical_contract import (
    HistoricalContractError,
    canonical_json_bytes,
    sha256_bytes,
    strict_json_bytes,
)


SEMANTIC_RATE_FIELDS = (
    "dataset",
    "provider",
    "product_id",
    "product_key",
    "rate_family",
    "rate_type",
    "application_type",
    "application_frequency",
    "calculation_frequency",
    "repayment_type",
    "loan_purpose",
    "term",
    "security_purpose",
    "ribbon_repayment_type",
    "lvr_tier",
    "lvr_source",
    "ribbon_rate_structure",
    "ribbon_fixed_term",
    "account_type",
    "ribbon_deposit_kind",
    "balance_min",
    "balance_max",
    "term_months",
    "interest_payment",
    "feature_set",
    "taxonomy_path",
    "account_class",
)
RATE_VALUE_FIELDS = ("rate", "comparison_rate")
DECIMAL_TEXT_FIELDS = {
    "rate",
    "comparison_rate",
    "balance_min",
    "balance_max",
    "term_months",
}
BOOLEAN_FIELDS = {"is_tailored", "ribbon_normalized"}
TEXT_TERM_RE = re.compile(
    r"(?<![A-Za-z0-9])\d+(?:\.\d+)?\s*"
    r"(?:d|day|days|w|wk|wks|week|weeks|m|mo|mos|mth|mths|month|months|"
    r"y|yr|yrs|year|years)(?![A-Za-z])",
    re.IGNORECASE,
)
COMPOUND_ISO_RE = re.compile(
    r"^P(?=\d)(?:(?:\d+(?:\.\d+)?)Y)(?:(?:\d+(?:\.\d+)?)M)"
    r"(?:(?:\d+(?:\.\d+)?)W)?(?:(?:\d+(?:\.\d+)?)D)?$",
    re.IGNORECASE,
)
SIMPLE_ISO_RE = re.compile(r"^P\d+(?:\.\d+)?[DWMY]$", re.IGNORECASE)
STRUCTURED_RANGE_RE = re.compile(
    r"^P\d+(?:\.\d+)?[DWMY]\s*/\s*P\d+(?:\.\d+)?[DWMY]$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class CanonicalRow:
    original_index: int
    digest: str
    payload: bytes


@dataclass(frozen=True)
class ParityResult:
    equal: bool
    left_count: int
    right_count: int
    missing: int
    extra: int


@dataclass(frozen=True)
class RawSemanticCollisionSummary:
    conflicting_groups: int
    conflicting_rows: int
    duplicate_same_value_groups: int
    duplicate_same_value_rows: int
    nonunique_rows: int
    records: tuple[Mapping[str, Any], ...]


def _decimal_text(value: float | Decimal) -> str:
    try:
        decimal = Decimal(str(value))
    except InvalidOperation as error:
        raise HistoricalContractError(f"invalid numeric value: {value!r}") from error
    if not decimal.is_finite():
        raise HistoricalContractError(f"non-finite numeric value: {value!r}")
    normalized = format(decimal.normalize(), "f")
    return "0" if normalized in {"-0", ""} else normalized


def canonical_scalar(value: Any) -> Any:
    if isinstance(value, bool) or value is None or isinstance(value, str):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, (float, Decimal)):
        return {"$decimal": _decimal_text(value)}
    raise HistoricalContractError(f"unsupported row scalar: {type(value).__name__}")


def canonical_value(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): canonical_value(child) for key, child in sorted(value.items())}
    if isinstance(value, (list, tuple)):
        return [canonical_value(child) for child in value]
    return canonical_scalar(value)


def _parse_embedded_json(value: Any) -> Any:
    if not isinstance(value, str) or not value or value[0] not in "[{":
        return value
    try:
        return strict_json_bytes(value.encode("utf-8"), source="embedded JSON")
    except HistoricalContractError:
        return value


def canonical_rows(
    rows: Iterable[Mapping[str, Any]],
    *,
    fields: Sequence[str] | None = None,
    text_fields: frozenset[str] = frozenset(),
) -> tuple[CanonicalRow, ...]:
    result: list[CanonicalRow] = []
    for index, source in enumerate(rows):
        selected = fields if fields is not None else tuple(sorted(source))
        row = {}
        for key in selected:
            raw = _parse_embedded_json(source.get(key))
            if key in BOOLEAN_FIELDS and raw in {0, 1, "0", "1", False, True}:
                raw = bool(int(raw))
            elif key in text_fields and not isinstance(raw, (dict, list, bool)):
                if raw is None:
                    raw = ""
                elif key in DECIMAL_TEXT_FIELDS and raw != "":
                    try:
                        raw = _decimal_text(Decimal(str(raw)))
                    except (InvalidOperation, ValueError):
                        raw = str(raw)
                elif not isinstance(raw, str):
                    raw = str(raw)
            row[key] = canonical_value(raw)
        payload = canonical_json_bytes(row)
        result.append(CanonicalRow(index, sha256_bytes(payload), payload))
    return tuple(result)


def compare_row_multisets(
    left: Iterable[Mapping[str, Any]],
    right: Iterable[Mapping[str, Any]],
    *,
    fields: Sequence[str] | None = None,
    text_fields: frozenset[str] = frozenset(),
) -> ParityResult:
    left_rows = canonical_rows(left, fields=fields, text_fields=text_fields)
    right_rows = canonical_rows(right, fields=fields, text_fields=text_fields)
    left_counter = Counter(row.payload for row in left_rows)
    right_counter = Counter(row.payload for row in right_rows)
    missing = sum((left_counter - right_counter).values())
    extra = sum((right_counter - left_counter).values())
    return ParityResult(not missing and not extra, len(left_rows), len(right_rows), missing, extra)


def text_fields_for(rows: Sequence[Mapping[str, Any]], fields: Sequence[str]) -> frozenset[str]:
    result = set()
    for field in fields:
        observed = {type(row.get(field)) for row in rows if row.get(field) not in {None, ""}}
        if observed and observed <= {str}:
            result.add(field)
    return frozenset(result)


def xlsx_rows(path: Any, sheet: str) -> Iterator[Mapping[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        iterator = worksheet.iter_rows(values_only=True)
        headers = tuple(str(value) for value in next(iterator))
        if len(headers) != len(set(headers)):
            raise HistoricalContractError(f"duplicate XLSX header in {sheet}")
        for values in iterator:
            yield dict(zip(headers, values))
    finally:
        workbook.close()


def sqlite_rows(connection: Any, table: str) -> Iterator[Mapping[str, Any]]:
    if table not in {"bank_products", "bank_rates"}:
        raise HistoricalContractError(f"SQLite table not allowlisted: {table}")
    cursor = connection.execute(f'SELECT * FROM "{table}" ORDER BY rowid')
    columns = tuple(item[0] for item in cursor.description)
    for values in cursor:
        yield dict(zip(columns, values))


def semantic_rate_collisions(
    rows: Iterable[Mapping[str, Any]],
) -> tuple[int, int, tuple[tuple[int, ...], ...]]:
    groups: dict[tuple[Any, ...], list[tuple[int, tuple[Any, ...]]]] = defaultdict(list)
    for index, row in enumerate(rows):
        identity = tuple(_hashable(row.get(field)) for field in SEMANTIC_RATE_FIELDS)
        values = tuple(_hashable(row.get(field)) for field in RATE_VALUE_FIELDS)
        groups[identity].append((index, values))
    collisions: list[tuple[int, ...]] = []
    for group in groups.values():
        if len(group) > 1 and len({values for _, values in group}) > 1:
            collisions.append(tuple(index for index, _ in group))
    collisions.sort()
    return len(collisions), sum(len(group) for group in collisions), tuple(collisions)


def semantic_collision_records(
    rows: Iterable[Mapping[str, Any]],
) -> dict[str, tuple[str, ...]]:
    """Return portable collision identities and distinct observed value rows.

    Repeated dates carry the same ambiguous tier populations. The locked corpus
    total is a union of those identities and distinct value rows, not a sum of
    the same ambiguity on every observation date.
    """
    groups: dict[tuple[Any, ...], set[tuple[Any, ...]]] = defaultdict(set)
    for row in rows:
        identity = tuple(_hashable(row.get(field)) for field in SEMANTIC_RATE_FIELDS)
        values = tuple(_hashable(row.get(field)) for field in RATE_VALUE_FIELDS)
        groups[identity].add(values)
    result: dict[str, tuple[str, ...]] = {}
    for identity, values in groups.items():
        if len(values) > 1:
            identity_bytes = canonical_json_bytes(list(identity))
            result[sha256_bytes(identity_bytes)] = tuple(
                sorted(sha256_bytes(canonical_json_bytes(list(value))) for value in values)
            )
    return result


def raw_semantic_collisions(
    products: Iterable[Mapping[str, Any]],
    flattened_rates: Iterable[Mapping[str, Any]],
) -> RawSemanticCollisionSummary:
    """Reproduce the forensic raw-tier census without conflating duplicates.

    Array position is used only to recover the embedded cleaned rate object.
    Identity is provider/product/family plus every raw semantic field except the
    advertised and comparison values. Same-value duplicates are reported
    separately from value-conflicting tiers.
    """
    rate_rows = tuple(flattened_rates)
    if not rate_rows:
        return RawSemanticCollisionSummary(0, 0, 0, 0, 0, ())

    product_map: dict[tuple[str, str], Mapping[str, Any]] = {}
    for product in products:
        identity = (
            str(product.get("provider") or ""),
            str(product.get("product_id") or ""),
        )
        details_raw = product.get("details_json")
        if not all(identity) or not isinstance(details_raw, str):
            raise HistoricalContractError("product cannot be mapped to embedded rate evidence")
        details = strict_json_bytes(
            details_raw.encode("utf-8"), source=f"{identity!r} details_json"
        )
        if not isinstance(details, Mapping):
            raise HistoricalContractError(f"product details are not an object: {identity!r}")
        prior = product_map.get(identity)
        if prior is not None and canonical_json_bytes(prior) != canonical_json_bytes(details):
            raise HistoricalContractError(
                f"duplicate provider/product identity has different details: {identity!r}"
            )
        product_map[identity] = details

    groups: dict[
        tuple[str, str, str, bytes],
        list[tuple[int, int, tuple[Any, Any]]],
    ] = defaultdict(list)
    for flattened_index, row in enumerate(rate_rows):
        identity = (str(row.get("provider") or ""), str(row.get("product_id") or ""))
        details = product_map.get(identity)
        if details is None:
            raise HistoricalContractError(f"flattened rate has no product evidence: {identity!r}")
        family = str(row.get("rate_family") or "").casefold()
        source_name = {"deposit": "depositRates", "lending": "lendingRates"}.get(family)
        if source_name is None:
            raise HistoricalContractError(f"unsupported flattened rate family: {family!r}")
        raw_rates = details.get(source_name, [])
        if not isinstance(raw_rates, list):
            raise HistoricalContractError(f"{identity!r} {source_name} is not an array")
        try:
            raw_index = int(row.get("rate_index"))
            raw = raw_rates[raw_index - 1]
        except (TypeError, ValueError, IndexError) as error:
            raise HistoricalContractError(
                f"flattened rate index does not map to embedded evidence: {identity!r}"
            ) from error
        if raw_index < 1 or not isinstance(raw, Mapping):
            raise HistoricalContractError(
                f"flattened rate index does not map to an object: {identity!r}"
            )
        semantics = {
            name: value
            for name, value in raw.items()
            if name not in {"rate", "comparisonRate"}
        }
        values = (canonical_value(raw.get("rate")), canonical_value(raw.get("comparisonRate")))
        groups[(*identity, family, canonical_json_bytes(semantics))].append(
            (flattened_index, raw_index, values)
        )

    records: list[Mapping[str, Any]] = []
    conflicting_groups = conflicting_rows = 0
    duplicate_groups = duplicate_rows_count = 0
    for (provider, product_id, family, semantics), members in groups.items():
        if len(members) < 2:
            continue
        values = {canonical_json_bytes(list(value)) for _, _, value in members}
        conflict = len(values) > 1
        if conflict:
            conflicting_groups += 1
            conflicting_rows += len(members)
        else:
            duplicate_groups += 1
            duplicate_rows_count += len(members)
        records.append(
            {
                "provider": provider,
                "product_id": product_id,
                "rate_family": family,
                "semantic_sha256": sha256_bytes(semantics),
                "classification": "value_conflict" if conflict else "same_value_duplicate",
                "flattened_indices": [index for index, _, _ in members],
                "raw_rate_indices": [index for _, index, _ in members],
                "row_count": len(members),
            }
        )
    records.sort(
        key=lambda item: (
            item["provider"], item["product_id"], item["rate_family"], item["semantic_sha256"]
        )
    )
    return RawSemanticCollisionSummary(
        conflicting_groups=conflicting_groups,
        conflicting_rows=conflicting_rows,
        duplicate_same_value_groups=duplicate_groups,
        duplicate_same_value_rows=duplicate_rows_count,
        nonunique_rows=conflicting_rows + duplicate_rows_count,
        records=tuple(records),
    )


def _hashable(value: Any) -> Any:
    canonical = canonical_value(value)
    if isinstance(canonical, (dict, list)):
        return canonical_json_bytes(canonical).decode("utf-8").rstrip("\n")
    return canonical


def _details_text(row: Mapping[str, Any]) -> str:
    details = _parse_embedded_json(row.get("details_json", ""))
    fragments = [
        str(row.get("product_name") or ""),
        str(row.get("description") or ""),
        str(row.get("rate_type") or ""),
        str(row.get("term") or ""),
    ]
    if isinstance(details, Mapping):
        # Frequencies describe accrual/payment cadence and are intentionally
        # excluded: they are not deposit-term evidence.
        for key, value in details.items():
            if str(key).casefold() not in {
                "applicationfrequency",
                "calculationfrequency",
            }:
                fragments.append(str(value))
    return " ".join(fragments)


def td_fallback_evidence(row: Mapping[str, Any]) -> str | None:
    if str(row.get("term_months") or "") != "12":
        return None
    term = str(row.get("term") or "").strip()
    if COMPOUND_ISO_RE.fullmatch(term):
        return "exact_iso"
    if SIMPLE_ISO_RE.fullmatch(term):
        # Ordinary P12M/P1Y observations are not fallbacks. Sub-month ISO rows
        # are exact-but-non-month evidence and are quarantined separately.
        number = float(term[1:-1])
        unit = term[-1:].upper()
        return "submonth_iso_quarantine" if (unit == "D" and number < 30) or (unit == "W" and number < 4) else None
    if STRUCTURED_RANGE_RE.fullmatch(term):
        return "structured_range"
    if "-" in term and SIMPLE_ISO_RE.search(term.split("-", 1)[0].strip()):
        return "text_derived"
    if TEXT_TERM_RE.search(_details_text(row)):
        return "text_derived"
    return "no_evidence"


def td_fallback_strata(rows: Iterable[Mapping[str, Any]]) -> Counter[str]:
    result: Counter[str] = Counter()
    for row in rows:
        evidence = td_fallback_evidence(row)
        if evidence is not None:
            result[evidence] += 1
    return result


def typed_rate_value(row: Mapping[str, Any], *, mixed_scale: bool = False) -> dict[str, Any]:
    raw = row.get("rate")
    result = {
        "legacy_raw_value": raw,
        "legacy_normalized_value": raw,
        "typed_value": None,
        "typed_unit": None,
        "unit_basis": None,
        "unit_status": "unavailable",
        "normalization_version": "historical-unit-v1",
    }
    try:
        numeric = Decimal(str(raw)) if not isinstance(raw, bool) else None
    except InvalidOperation:
        numeric = None
    if mixed_scale or numeric is None or not numeric.is_finite():
        result["unit_status"] = "quarantined_mixed_scale" if mixed_scale else "unavailable"
        return result
    magnitude = abs(numeric)
    typed = raw if isinstance(raw, (int, float)) else _decimal_text(numeric)
    if magnitude <= 0.2:
        result.update(typed_value=typed, typed_unit="fraction", unit_basis="magnitude_proven", unit_status="derived")
    elif magnitude > 1 and magnitude <= 100:
        result.update(typed_value=typed, typed_unit="percentage_points", unit_basis="magnitude_proven", unit_status="derived")
    else:
        result["unit_status"] = "quarantined_ambiguous_scale"
    return result


def classify_savings_product(row: Mapping[str, Any]) -> dict[str, str]:
    dataset = str(row.get("dataset") or "").upper()
    category = str(row.get("category") or "").upper()
    text = " ".join(
        str(row.get(key) or "")
        for key in ("product_name", "description", "account_type", "taxonomy_path")
    ).casefold()
    if dataset == "TD" or category == "TERM_DEPOSITS" or "term deposit" in text:
        return {"class": "term_deposit", "status": "confirmed_exclusion", "basis": "structured_td"}
    if dataset == "BUSINESS_LOANS" or "business loan" in text:
        return {"class": "business", "status": "confirmed_exclusion", "basis": "structured_business"}
    if any(token in text for token in ("offset", "transaction account", "everyday account", "cash management")):
        return {"class": "non_savings_account", "status": "quarantined", "basis": "name_requires_structured_confirmation"}
    if any(token in text for token in ("restricted", "staff only", "private client")):
        return {"class": "restricted", "status": "quarantined", "basis": "eligibility_incomplete"}
    if dataset == "SAVINGS" and category == "TRANS_AND_SAVINGS_ACCOUNTS":
        return {"class": "ordinary_savings", "status": "ambiguous", "basis": "legacy_bucket_only"}
    return {"class": "unknown", "status": "quarantined", "basis": "insufficient_evidence"}


def absence_evidence(rows: Sequence[Mapping[str, Any]], kind: str) -> Mapping[str, Any]:
    if rows:
        return {"state": "observed_cleaned_projection", "count": len(rows)}
    return {
        "state": "unavailable",
        "value": None,
        "reason": f"no retained {kind} rows; absence does not prove zero or unrestricted",
    }


def duplicate_rows(rows: Iterable[Mapping[str, Any]]) -> tuple[tuple[int, ...], ...]:
    groups: dict[bytes, list[int]] = defaultdict(list)
    for row in canonical_rows(rows):
        groups[row.payload].append(row.original_index)
    return tuple(tuple(indices) for indices in groups.values() if len(indices) > 1)
