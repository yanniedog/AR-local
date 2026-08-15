from __future__ import annotations

import tracemalloc
from pathlib import Path

import openpyxl

from cdr_xlsx import write_workbook


class GeneratedRows:
    def __init__(self, count: int, value_size: int) -> None:
        self.count = count
        self.value = "x" * value_size

    def __len__(self) -> int:
        return self.count

    def __iter__(self):
        for index in range(self.count):
            yield {"index": index, "value": self.value}


def test_write_workbook_streams_rows_without_materializing_sheet_xml(tmp_path: Path) -> None:
    output = tmp_path / "streamed.xlsx"
    rows = GeneratedRows(count=8_000, value_size=1_000)

    tracemalloc.start()
    write_workbook(output, {"generated": rows})
    _, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()

    assert peak < 4_000_000
    workbook = openpyxl.load_workbook(output, read_only=True)
    sheet = workbook["generated"]
    assert sum(1 for _ in sheet.iter_rows()) == 8_001
    assert sheet["A2"].value == 0
    assert sheet["B8001"].value == "x" * 1_000


def test_write_workbook_preserves_empty_sheet(tmp_path: Path) -> None:
    output = tmp_path / "empty.xlsx"
    write_workbook(output, {"empty": []})

    workbook = openpyxl.load_workbook(output, read_only=True)
    assert workbook["empty"]["A1"].value == "empty"
