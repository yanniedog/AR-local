import json
import sqlite3
from copy import deepcopy
from pathlib import Path

import app_payload_mobile
import cdr_outputs
import openpyxl
from cdr_product_facts import audit_records, clean_fact_rows, compact_facts, extract_product_facts


FIXTURE = Path(__file__).parent / "fixtures" / "product_facts_real_2026-05-19.json"


def captured():
    return json.loads(FIXTURE.read_text(encoding="utf-8"))["products"]


def test_real_cdr_facts_are_typed_unique_grouped_and_currency_safe():
    products = captured()
    up = products[0]["record"]
    facts = extract_product_facts(up, "Mortgage|Up|up-home")
    assert len({fact["fact_id"] for fact in facts}) == len(facts)
    assert next(fact for fact in facts if fact["canonical_key"] == "rate.advertised")["value"] == 0.057
    max_lvr = next(fact for fact in facts if fact["canonical_key"] == "constraint.value" and fact["unit"] == "fraction")
    assert max_lvr["value"] == 0.9
    min_age = next(fact for fact in facts if fact["canonical_key"] == "eligibility.value")
    assert (min_age["value"], min_age["unit"]) == (18.0, "year")
    usd = extract_product_facts(products[1]["record"], "Savings|Bank Australia|basic")
    amount = next(fact for fact in usd if fact["canonical_key"] == "fee.amount")
    assert (amount["value"], amount["unit"]) == (175.0, "USD")


def test_real_tier_emits_explicit_fraction_range_and_preserves_leaves():
    facts = extract_product_facts(captured()[2]["record"], "Mortgage|BOQS|basic")
    range_fact = next(fact for fact in facts if fact["value_type"] == "range")
    assert (range_fact["kind"], range_fact["min_value"], range_fact["max_value"], range_fact["unit"]) == (
        "tier", 0.6001, 0.7, "fraction",
    )
    assert any(fact["canonical_key"] == "tier.minimum" for fact in facts)
    compact = compact_facts(captured()[2]["record"], "Mortgage|BOQS|basic")
    assert any(fact.get("minValue") == 0.6001 and fact.get("maxValue") == 0.7 for fact in compact)


def test_non_numeric_rate_and_range_values_are_preserved_without_invalid_sql_types():
    record = {
        "productId": "non-numeric",
        "lendingRates": [{
            "rate": "POA",
            "tiers": [{
                "minimumValue": "Varies",
                "maximumValue": "N/A",
                "unitOfMeasure": "PERCENT",
            }],
        }],
    }
    base = {
        "dataset": "Mortgage", "provider": "Example", "product_id": "non-numeric",
        "product_key": "Mortgage|Example|non-numeric", "product_name": "Example loan",
    }
    facts = extract_product_facts(record, base["product_key"])
    advertised = next(fact for fact in facts if fact["canonical_key"] == "rate.advertised")
    assert (advertised["value_type"], advertised["value"], advertised["unit"]) == ("text", "POA", "text")
    assert not any(fact["value_type"] == "range" for fact in facts)

    rows = [{**row, "run_date": "2026-05-19"} for row in clean_fact_rows(record, base)]
    with sqlite3.connect(":memory:") as con:
        cdr_outputs.ensure_db(con)
        cdr_outputs.insert_rows(con, "bank_product_facts", rows)
        assert con.execute("SELECT COUNT(*) FROM bank_product_facts").fetchone()[0] == len(rows)


def test_numeric_product_id_is_preserved_as_opaque_text():
    facts = extract_product_facts({"productId": "001234567890123456789"}, "Mortgage|Bank|stable")
    product_id = next(fact for fact in facts if fact["canonical_key"] == "product.id")
    assert (product_id["value_type"], product_id["value"], product_id["unit"]) == (
        "text", "001234567890123456789", "text",
    )


def test_percent_style_rates_are_normalized_to_fractions_everywhere():
    record = {
        "productId": "percent-rates",
        "lendingRates": [{"lendingRateType": "VARIABLE", "rate": "5.0", "comparisonRate": "5.25"}],
    }
    facts = extract_product_facts(record, "Mortgage|Bank|percent-rates")
    assert next(fact for fact in facts if fact["canonical_key"] == "rate.advertised")["value"] == 0.05
    assert next(fact for fact in facts if fact["canonical_key"] == "rate.comparison")["value"] == 0.0525
    compact = next(fact for fact in compact_facts(record, "Mortgage|Bank|percent-rates") if fact["kind"] == "rate")
    assert (compact["value"], compact["comparisonValue"]) == (0.05, 0.0525)


def test_variable_zero_fee_is_unpublished_not_free():
    fixture = json.loads(
        (Path(__file__).parent / "fixtures" / "greater_bank_fee_details_2026-05-19.json").read_text(encoding="utf-8")
    )
    record = fixture["record"]
    facts = extract_product_facts(record, "Mortgage|Greater Bank|loan")
    pexa = next(
        fact for fact in facts
        if fact["canonical_key"] == "fee.amount" and fact["qualifiers"].get("feeType") == "VARIABLE"
    )
    assert (pexa["value_type"], pexa["value"], pexa["mapping"]) == ("text", "unpublished", "canonical")
    compact = next(fact for fact in compact_facts(record, "Mortgage|Greater Bank|loan") if fact["label"].startswith("Property Exchange"))
    assert "value" not in compact


def test_negative_text_rules_take_precedence_over_positive_substrings():
    record = {
        "description": (
            "No offset account is available. Customers cannot redraw and cannot make extra repayments."
        ),
    }
    facts = extract_product_facts(record, "Mortgage|Bank|stable")
    tagged = {
        (fact["canonical_key"], fact["value"])
        for fact in facts if fact["mapping"] == "canonical_text"
    }
    assert ("feature.offset", False) in tagged
    assert ("feature.redraw", False) in tagged
    assert ("feature.extra_repayments", False) in tagged
    assert not any(value is True for key, value in tagged if key.startswith("feature."))


def test_compact_facts_consolidate_structural_fields_and_keep_exact_conditions():
    compact = compact_facts(captured()[0]["record"], "Mortgage|Up|up-home")
    assert not any(fact["canonicalKey"] in {"feature.type", "currency", "fee.method"} for fact in compact)
    offset = next(fact for fact in compact if fact["canonicalKey"] == "feature.offset" and fact.get("sourceType") == "OFFSET")
    assert offset["label"] == "Offset"
    assert offset["condition"] == "Use your Up spending account and Savers as offsets."
    assert offset["groupId"]
    assert "source_path" not in json.dumps(compact)


def test_audit_proves_every_nonempty_scalar_covered_and_reports_unmatched_text():
    records = [(row["record"]["productId"], row["record"]) for row in captured()]
    report = audit_records(records)
    assert report["unmapped_nonempty_scalar_paths"] == []
    assert report["duplicate_fact_ids"] == []
    assert report["covered_nonempty_scalars"] == report["observed_nonempty_scalars"]
    assert report["text_coverage"]["unmatched"] > 0
    assert report["text_coverage"]["unmatched_semantic_status"] == "preserved_not_equivalent"


def test_semantic_ids_survive_reordering_and_repeated_tiers_remain_unique():
    record = deepcopy(captured()[0]["record"])
    before = extract_product_facts(record, "stable-product")
    record["features"].reverse()
    record["lendingRates"].reverse()
    after = extract_product_facts(record, "stable-product")
    signature = lambda fact: (
        fact["canonical_key"], fact["source_value_json"], fact["mapping"],
        fact["qualifiers"].get("lendingRateType"), fact["qualifiers"].get("featureType"),
    )
    assert {signature(fact): fact["fact_id"] for fact in before} == {
        signature(fact): fact["fact_id"] for fact in after
    }
    tiered = deepcopy(captured()[2]["record"])
    tiered["lendingRates"][0]["tiers"].append(deepcopy(tiered["lendingRates"][0]["tiers"][0]))
    tier_facts = [fact for fact in extract_product_facts(tiered, "repeated-tier") if fact["value_type"] == "range"]
    assert len(tier_facts) == 2
    assert len({fact["fact_id"] for fact in tier_facts}) == 2


def test_duplicate_entity_ids_survive_cardinality_changes():
    record = {
        "productId": "fees", "name": "Fees",
        "fees": [
            {"name": "Service fee", "feeType": "PERIODIC", "amount": "5"},
            {"name": "Service fee", "feeType": "PERIODIC", "amount": "10"},
        ],
    }
    full_before = [fact for fact in extract_product_facts(record, "stable-fees") if fact["canonical_key"] == "fee.amount"]
    compact_before = [fact for fact in compact_facts(record, "stable-fees") if fact["kind"] == "fee"]
    record["fees"].pop(0)
    full_after = [fact for fact in extract_product_facts(record, "stable-fees") if fact["canonical_key"] == "fee.amount"]
    compact_after = [fact for fact in compact_facts(record, "stable-fees") if fact["kind"] == "fee"]
    assert next(fact["fact_id"] for fact in full_before if fact["value"] == 10) == full_after[0]["fact_id"]
    assert next(fact["id"] for fact in compact_before if fact.get("value") == 10) == compact_after[0]["id"]


def test_text_taxonomy_scopes_negation_to_each_clause():
    facts = compact_facts({
        "productId": "variants", "name": "Variants",
        "description": "No redraw for fixed loans; redraw facility is available for variable loans.",
    }, "variants")
    redraw = [fact for fact in facts if fact["canonicalKey"] == "feature.redraw"]
    assert {fact["value"] for fact in redraw} == {False, True}
    assert {fact["condition"] for fact in redraw} == {
        "No redraw for fixed loans", "redraw facility is available for variable loans.",
    }


def test_fee_free_wording_does_not_negate_available_features():
    facts = compact_facts({
        "productId": "fee-free", "name": "Fee-free features",
        "features": [
            {"featureType": "REDRAW", "additionalInfo": "No redraw fee"},
            {"featureType": "EXTRA_REPAYMENTS", "additionalInfo": "No extra repayment fee"},
        ],
    }, "fee-free")
    assert {fact["value"] for fact in facts if fact["canonicalKey"] == "feature.redraw"} == {True}
    assert {fact["value"] for fact in facts if fact["canonicalKey"] == "feature.extra_repayments"} == {True}


def test_product_fields_are_root_scoped_and_nested_names_keep_entity_kind():
    facts = extract_product_facts({
        "productId": "root", "name": "Root product", "description": "Root description",
        "fees": [{"name": "Valuation fee", "description": "At cost", "feeType": "EVENT"}],
        "features": [{"name": "Redraw", "description": "Available", "featureType": "REDRAW"}],
    }, "root")
    product_names = [fact["value"] for fact in facts if fact["canonical_key"] == "product.name"]
    product_descriptions = [fact["value"] for fact in facts if fact["canonical_key"] == "product.description"]
    assert product_names == ["Root product"]
    assert product_descriptions == ["Root description"]
    assert {fact["canonical_key"] for fact in facts} >= {
        "fee.name", "fee.description", "feature.name", "feature.description",
    }


def test_compact_facts_preserve_legacy_fee_rates_and_applicability_values():
    facts = compact_facts({
        "productId": "priced", "name": "Priced",
        "fees": [{
            "name": "Usage fee", "feeType": "TRANSACTION",
            "balanceRate": "1", "transactionRate": "2", "accruedRate": "3",
        }],
        "lendingRates": [{
            "lendingRateType": "VARIABLE", "rate": "5.5",
            "applicabilityConditions": [{
                "rateApplicabilityType": "MINIMUM_BALANCE", "additionalValue": "5000",
            }],
        }],
    }, "priced")
    fee_rates = [fact for fact in facts if fact["kind"] == "fee" and fact.get("unit") == "fraction"]
    assert {fact["value"] for fact in fee_rates} == {0.01, 0.02, 0.03}
    minimum = next(fact for fact in facts if fact["canonicalKey"] == "condition.minimum_balance")
    assert minimum["value"] == 5000
    assert minimum["unit"] == "AUD"


def test_fact_ids_use_stable_product_identity_not_mutable_name_or_category():
    record = {
        "productId": "stable-id", "name": "Original name", "productCategory": "RESIDENTIAL_MORTGAGES",
        "features": [{"featureType": "OFFSET", "additionalValue": "Available"}],
    }
    before = clean_fact_rows(record, {
        "dataset": "Mortgage", "provider": "Example Bank", "product_id": "stable-id",
        "product_key": "Example Bank|stable-id|RESIDENTIAL_MORTGAGES|Original name",
        "product_name": "Original name",
    })
    renamed = {**record, "name": "Renamed product", "productCategory": "TRANS_AND_SAVINGS_ACCOUNTS"}
    after = clean_fact_rows(renamed, {
        "dataset": "Mortgage", "provider": "Example Bank", "product_id": "stable-id",
        "product_key": "Example Bank|stable-id|TRANS_AND_SAVINGS_ACCOUNTS|Renamed product",
        "product_name": "Renamed product",
    })
    ids = lambda rows, canonical: next(row["fact_id"] for row in rows if row["canonical_key"] == canonical)
    assert ids(before, "product.id") == ids(after, "product.id")
    assert ids(before, "feature.type") == ids(after, "feature.type")


def test_audit_reports_source_failures_and_does_not_claim_complete_capture():
    report = audit_records(
        [("up-home", captured()[0]["record"])],
        [{"bank": "Example", "phase": "product_detail", "status": "403"}],
    )
    assert report["complete"] is False
    assert report["detail_failures"][0]["status"] == "403"
    assert report["source_failures"] == report["detail_failures"]


def test_search_index_uses_vetted_fact_values_labels_and_conditions_not_raw_paths_or_urls():
    facts = compact_facts(captured()[0]["record"], "key")
    index = app_payload_mobile.build_search_index([], {"key": {"facts": facts}}, run_date="2026-05-19")
    text = index["products"]["key"]
    assert "feature offset" in text
    assert "use your up spending account and savers as offsets" in text
    assert "source_path" not in text and "http" not in text


def test_search_index_ignores_scalar_search_terms_without_failing():
    index = app_payload_mobile.build_search_index(
        [], {"key": {"facts": [{"searchTerms": "offset account", "value": True}]}},
        run_date="2026-05-19",
    )
    assert index["products"]["key"] == "key"


def test_compact_payload_has_a_bounded_entity_count_and_size():
    all_facts = [fact for product in captured() for fact in compact_facts(product["record"], product["record"]["productId"])]
    encoded = json.dumps(all_facts, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    assert len(all_facts) < 100
    assert len(encoded) < 40_000


def test_sqlite_fact_constraints_and_filter_indexes_are_used():
    record = captured()[0]["record"]
    base = {"dataset": "Mortgage", "provider": "Up", "product_id": "up-home", "product_key": "Up|up-home", "product_name": "Up Home Loan"}
    rows = clean_fact_rows(record, base)
    with sqlite3.connect(":memory:") as con:
        cdr_outputs.ensure_db(con)
        cdr_outputs.insert_rows(con, "bank_product_facts", [{"run_date": "2026-05-19", **row} for row in rows])
        numeric = con.execute("EXPLAIN QUERY PLAN SELECT product_key FROM bank_product_facts WHERE run_date=? AND dataset=? AND canonical_key=? AND value_number>=?", ("2026-05-19", "Mortgage", "rate.advertised", 0.05)).fetchall()
        categorical = con.execute("EXPLAIN QUERY PLAN SELECT product_key FROM bank_product_facts WHERE run_date=? AND dataset=? AND canonical_key=? AND value_text=?", ("2026-05-19", "Mortgage", "loan.purpose", "OWNER_OCCUPIED")).fetchall()
        assert any("idx_bank_product_facts_numeric" in row[-1] for row in numeric)
        assert any("idx_bank_product_facts_categorical" in row[-1] for row in categorical)


def test_normal_outputs_include_facts_json_xlsx_and_sqlite(tmp_path: Path):
    run = tmp_path / "2026-05-19"
    for index, product in enumerate(captured()):
        path = run / "banks" / product["dataset"] / product["provider"] / f"product-{index}" / "id"
        path.mkdir(parents=True)
        (path / "product-detail.json").write_text(json.dumps({"data": product["record"]}), encoding="utf-8")
    out = tmp_path / "exports"
    cdr_outputs.build_outputs(run, out_dir=out)
    exported = json.loads((out / "banks-2026-05-19.json").read_text(encoding="utf-8"))
    assert exported["product_facts"]
    assert exported["product_change_summary"]["normalization_version"] == "cdr-product-facts-2"
    workbook = openpyxl.load_workbook(out / "banks-2026-05-19.xlsx", read_only=True)
    assert "product_facts" in workbook.sheetnames
    with sqlite3.connect(out / "local-cdr.sqlite") as con:
        assert con.execute("SELECT COUNT(*) FROM bank_product_facts").fetchone()[0] == len(exported["product_facts"])


def test_normal_outputs_index_previous_run_product_changes(tmp_path: Path):
    def write_run(day: str, description: str) -> Path:
        run = tmp_path / day
        path = run / "banks" / "Mortgage" / "Example Bank" / "Home Loan" / "stable-id"
        path.mkdir(parents=True)
        record = {
            "productId": "stable-id", "name": "Home Loan", "brand": "Example Bank",
            "description": description, "features": [], "eligibility": [], "constraints": [],
            "fees": [], "lendingRates": [],
        }
        (path / "product-detail.json").write_text(json.dumps({"data": record}), encoding="utf-8")
        return run

    previous = write_run("2026-05-18", "An offset account is available.")
    current = write_run("2026-05-19", "No offset account is available.")
    cdr_outputs.build_outputs(previous)
    cdr_outputs.build_outputs(current)
    exported = json.loads((current / "_exports" / "banks-2026-05-19.json").read_text(encoding="utf-8"))
    assert exported["product_change_summary"]["previous_run_date"] == "2026-05-18"
    assert exported["product_change_summary"]["change_count"] == len(exported["product_changes"])
    assert exported["product_changes"]
    workbook = openpyxl.load_workbook(current / "_exports" / "banks-2026-05-19.xlsx", read_only=True)
    assert {"product_changes", "change_summary"} <= set(workbook.sheetnames)
    with sqlite3.connect(current / "_exports" / "local-cdr.sqlite") as con:
        assert con.execute("SELECT COUNT(*) FROM bank_product_changes").fetchone()[0] == len(exported["product_changes"])
        plan = con.execute(
            "EXPLAIN QUERY PLAN SELECT product_id FROM bank_product_changes WHERE run_date=? AND provider=? AND event_type=? AND canonical_key=?",
            ("2026-05-19", "Example Bank", "condition_changed", "feature.offset"),
        ).fetchall()
        assert any("idx_bank_product_changes_lookup" in row[-1] for row in plan)

    current_detail = next((current / "banks").rglob("product-detail.json"))
    refreshed = json.loads(current_detail.read_text(encoding="utf-8"))
    refreshed["data"]["description"] = "Offset access now requires an eligible transaction account."
    current_detail.write_text(json.dumps(refreshed), encoding="utf-8")
    cdr_outputs.build_outputs(current)
    rebuilt = json.loads((current / "_exports" / "banks-2026-05-19.json").read_text(encoding="utf-8"))
    assert "eligible transaction account" in json.dumps(rebuilt["product_changes"])


def test_cross_day_source_paths_do_not_create_metadata_changes(tmp_path: Path):
    def write(day: str) -> Path:
        run = tmp_path / day
        detail = run / "banks" / "Mortgage" / "Example Bank" / "Home Loan" / "stable-id"
        detail.mkdir(parents=True)
        record = {"productId": "stable-id", "name": "Home Loan", "features": [{"featureType": "OFFSET"}]}
        (detail / "product-detail.json").write_text(json.dumps({"data": record}), encoding="utf-8")
        return run

    previous, current = write("2026-05-18"), write("2026-05-19")
    cdr_outputs.build_outputs(previous)
    cdr_outputs.build_outputs(current)
    exported = json.loads((current / "_exports" / "banks-2026-05-19.json").read_text(encoding="utf-8"))
    assert exported["product_changes"] == []


def test_failed_current_detail_does_not_report_product_removal(tmp_path: Path):
    previous = tmp_path / "2026-05-18"
    detail = previous / "banks" / "Mortgage" / "Example Bank" / "Home Loan" / "stable-id"
    detail.mkdir(parents=True)
    (detail / "product-detail.json").write_text(
        json.dumps({"data": {"productId": "stable-id", "name": "Home Loan", "features": []}}),
        encoding="utf-8",
    )
    current = tmp_path / "2026-05-19"
    banks = current / "banks"
    banks.mkdir(parents=True)
    (banks / "failures.jsonl").write_text(
        json.dumps({"phase": "product_detail", "bank": "Example Bank", "product_id": "stable-id", "status": 503}) + "\n",
        encoding="utf-8",
    )
    cdr_outputs.build_outputs(previous)
    cdr_outputs.build_outputs(current)
    exported = json.loads((current / "_exports" / "banks-2026-05-19.json").read_text(encoding="utf-8"))
    assert exported["product_changes"] == []
    assert exported["product_change_summary"]["suppressed_incomplete_products"] == 1
